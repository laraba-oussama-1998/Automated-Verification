import sys
if hasattr(sys, "_MEIPASS"): # if the script is started from an executable file
    with open("logs.txt", "w") as f_logs:
        sys.stdout = f_logs
        sys.stderr = f_logs
import eel
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import math
import traceback    
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import os
import shutil
from google.cloud import vision
import cv2 as cv
from matplotlib import pyplot as plt
from collections import OrderedDict
from pdf2image import convert_from_path
from collections import OrderedDict
import fitz
from pathlib import Path
from google.oauth2 import service_account
import json


# import sys
# logfile = open('program_output.txt', 'w')
# sys.stdout = logfile
# sys.stderr = logfile

eel.init('web') 

# Afficher une fenêtre contextuelle pour sélectionner le fichier PDF à traiter.
@eel.expose
def select_and_send_pdf_path():
    global pdf_path 
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    root.attributes('-topmost', 1)
    pdf_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])

    if pdf_path:
        return pdf_path
    else:
        return "" 

# Afficher une fenêtre contextuelle pour sélectionner le fichier Excel à traiter.
@eel.expose
def select_and_send_excel_path():
    global file_path 
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    root.attributes('-topmost', 1)
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])

    if file_path:
        return file_path
    else:
        return ""  

# Affiche une fenêtre contextuelle pour sélectionner le chemin de téléchargement du fichier.
@eel.expose
def select_and_send_folder_path():
    global folder_path 
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    root.attributes('-topmost', 1)
    folder_path = filedialog.askdirectory()

    if folder_path:
        return folder_path
    else:
        return ""  

# Déplacez le fichier résultat vers un chemin défini par l'utilisateur.
@eel.expose
def telecharger_fichier(fichier, download_path, folder_path):
    if (folder_path == ''):
        if os.path.exists(fichier):
            if os.path.exists(download_path+'/'+fichier):
                os.remove(download_path+'/'+fichier)
                shutil.copy(fichier, download_path)
            else :
                shutil.copy(fichier, download_path)
            return (download_path+'/'+fichier)
    else :
        if os.path.exists(fichier):
            if os.path.exists(folder_path+'/'+fichier):
                os.remove(folder_path+'/'+fichier)
                shutil.copy(fichier, folder_path)
            else :
                shutil.copy(fichier, folder_path)
            return(folder_path+'/'+fichier)

# L'utilisation de l'OCR de Google pour l'extraction de texte.
def ocr_call(image_path):
    image = cv.imread(image_path,cv.IMREAD_UNCHANGED)
    # Le fichier JSON qui contient les informations d'identification de l'OCR
    service_account_key_json = {
                "type": "service_account",
                "project_id": "dauntless-tube-368113",
                "private_key_id": "3d1cd3d7929069bb2faa0f49463f7be26bf09296",
                "private_key": os.getenv("MY_PRIVATE_KEY"),
                "client_email": "dauntless-tube-368113@appspot.gserviceaccount.com",
                "client_id": "104316379693227995461",
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/dauntless-tube-368113%40appspot.gserviceaccount.com",
                "universe_domain": "googleapis.com"
                }
    credentials = service_account.Credentials.from_service_account_info(service_account_key_json)

    client = vision.ImageAnnotatorClient(credentials=credentials)
    success, img_jpg = cv.imencode(".jpg", image)
    byte_img = img_jpg.tobytes()
    google_img = vision.Image(content=byte_img)
    response = client.document_text_detection(image=google_img)
    return response, image

# Récupérer le contenu de la réponse du service OCR Google.
def get_content(response, image):
    content_1 = {}
    content_2 = {}
    
    # verifier si la largeur d'image est superieur à la longueur càd c'est l'attestation
    if image.shape[0] < image.shape[1]:
        for obj in response.text_annotations[1:]:
            description = obj.description
            vertices = obj.bounding_poly.vertices
            if (vertices[0].x+vertices[1].x)//2 < image.shape[1]//2:
                content_1[((vertices[0].x+vertices[1].x)//2,(vertices[0].y+vertices[3].y)//2)] = description
            else:
                content_2[((vertices[0].x+vertices[1].x)//2,(vertices[0].y+vertices[3].y)//2)] = description
        # Les dictionnaires sont ordonnés du plus haut au plus bas.
        # content_1 a le contenu de l'extraction de la moitié gauche de l'image.
        # content_2 a le contenu de l'extraction de la moitié droite de l'image.
        content_1 = OrderedDict(sorted(content_1.items(), key=lambda x:x[0][1]))
        content_2 = OrderedDict(sorted(content_2.items(), key=lambda x:x[0][1]))
        
        return content_1, content_2
    else:
        #Ici, nous avons uniquement content_1 car nous traitons le relevé ici.
        for obj in response.text_annotations[1:]:
            description = obj.description
            vertices = obj.bounding_poly.vertices
            content_1[((vertices[0].x+vertices[1].x)//2,(vertices[0].y+vertices[3].y)//2)] = description
        
        content_1 = OrderedDict(sorted(content_1.items(), key=lambda x:x[0][1]))
    
        return content_1, None

# Organiser les résultats de l'extraction ligne par ligne.
def get_lines(content):
    keys = list(content.keys())
    lines_dict = {'line1':[keys[0]]}
    lines = [keys[0]]
    index = 1
    current_line = 0
    while index < len(keys):
        x, y = keys[index]
        #Les éléments qui sont distants de moins de 10 pixels appartiennent à la même ligne.
        if abs(lines[current_line][1]-y) < 10:
            x1 = (lines[current_line][0] + x) // 2
            y1 = (lines[current_line][1] + y) // 2
            lines[current_line] = (x1, y1)
            index += 1
            lines_dict[f'line{current_line+1}'].append((x, y))
        else:
            #sinon on va creer une nouvelle ligne
            lines.append(keys[index])
            lines_dict[f'line{current_line+1}'] = sorted(lines_dict[f'line{current_line+1}'],key= lambda x: x[0])
            current_line +=1
            lines_dict[f'line{current_line+1}'] = [keys[index]]
            index+=1
            
    return lines, lines_dict

# Trouver le numéro de CCP sur l'attestation.
def find_ccp(lines_dict, content, image):
    ccp_numbers = {}
    # Lire le contenu extrait de manière organisée, ligne par ligne.
    for line, coordinations in lines_dict.items():
        line_content = ''
        coord_first_digit = (0,0)
        for coord in coordinations:
            if len(re.findall('[0-9]',content.get(coord))) > 0:
                if coord == (0,0):
                    coord_first_digit = coord
                line_content += content.get(coord)+' '#+str(coord)+'  '
        # Vérifiez si la ligne contient plus de 5 chiffres, c'est-à-dire s'il s'agit du premier numéro de CCP
        total_digits = re.findall('[0-9]',line_content)
        if len(total_digits) > 5 and coord[1] < image.shape[0]//2:
            
            if ccp_numbers.keys():
                first_ccp_key = list(ccp_numbers.keys())[0]
                if (coord[1] - first_ccp_key[1]) > 100:
                    # print('break')
                    break
                else:
                    ccp_numbers[coord] = ''.join(total_digits)
            else:
                ccp_numbers[coord] = ''.join(total_digits)
        
        if len(ccp_numbers) == 2:
            break
    return ccp_numbers

# l'encadration de l'image
def crop_images(img):
    image = np.array(img)
    base_image = image.copy()
    
    gray = cv.cvtColor(image, cv.COLOR_BGR2GRAY)
    blur = cv.GaussianBlur(gray, (7,7), 0)
    thresh = cv.threshold(blur, 0, 255, cv.THRESH_BINARY_INV + cv.THRESH_OTSU)[1]
    kernal = cv.getStructuringElement(cv.MORPH_RECT, (3, 13))
    dilate = cv.dilate(thresh, kernal, iterations=1)

    cnts = cv.findContours(dilate, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]
    cnts = sorted(cnts, key=lambda x: cv.boundingRect(x)[0])

    # Créer un cadre englobant tous les cadres détectés
    x, y, w, h = 0, 0, 0, 0
    for c in cnts:
        cx, cy, cw, ch = cv.boundingRect(c)
        cv.rectangle(image, (cx, cy), (cx+cw, cy+ch), (36, 255, 12), 2)
        if x == 0 and y == 0 and w == 0 and h == 0:
            x, y, w, h = cx, cy, cw, ch
        else:
            x = min(x, cx)
            y = min(y, cy)
            w = max(w, cx + cw - x)
            h = max(h, cy + ch - y)
    # Recadrer l'image en fonction des dimensions du cadre englobant
    image_cropped = base_image[y:y + h, x:x + w]
    return image_cropped


# Vérifiez si le relevé est renversé.
def is_flipped(image, lines_dict_1, content_1):
    items = list(lines_dict_1.items())
    flipped = False

    # Vérifiez si les 20 premiers éléments se trouvent en haut ou en bas de la page
    first_coord = sum([items[i][1][0][1] for i in range(5)])/5
    
    if (image.shape[0] - first_coord) < image.shape[0]//2:
        return True
    # Vérifier s'il existe un mot qui appartient à cette liste, car ces mots sont toujours au début du relevé
    # afin de détecter la direction du relevé
    # Si le mot est dans la première partie, alors la direction du relevé est normale
    for line in items[:int(len(items)*0.20)]:
        for elemenet in line[1]:
            if content_1.get(elemenet) in ['CENTRE','COMPTE','COURANT','POSTAL', 'RECONSTRUCT', 'CHEQUES','POSTAUX', 'POSTE', 'RELEVE', 'DESIGNATION', 'DATE']:
                return False
    # sinon le releve est renversé
    for line in items[int(len(items)*0.80):]:
        for elemenet in line[1]:
            if content_1.get(elemenet) in ['CENTRE','COMPTE','COURANT','POSTAL', 'RECONSTRUCT', 'CHEQUES','POSTAUX', 'POSTE', 'RELEVE', 'DESIGNATION', 'DATE']:
                return True

    return None

# Trouver le numéro du CCP sur le relevé.
def find_ccp_releve(flipped, lines_dict_1, content_1):
    # Trouver la ligne où le compte CCP existe.
    for line in lines_dict_1.items():
        text = ''
        for element in line[1]:
            text += content_1.get(element)+' '
        check = [1  if element in ['COMPTE','COURANT','POSTAL', 'NO', 'CLE','CLR', 'SUR'] else 0 for element in text.split()]
        if sum(check) > 2:
            break
    
    # extraire le numero de compte
    if flipped:
        ccp_list = re.findall('[0-9]+',text)
        return ccp_list
    else:
        ccp_list = re.findall('[0-9]+', ' '.join(text.split()[::-1]))
        return ccp_list

# Convertir les documents scannés en images.
def convert_pdf_to_images(pdf_path):
    images = convert_from_path(pdf_path)
    for i in range(0,len(images)):
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return 0
        # Sauvgardez les pages comme des image sur le pdf.
        cropped_image = crop_images(images[i])
        if i%2:
            cv.imwrite('releve_'+ str(i//2+1) +'.jpg', cropped_image)
        else:
            cv.imwrite('attestation_'+ str(i//2+1) +'.jpg', cropped_image)
    return len(images)


# Cette fonction main est responsable du regroupement des autres fonctions.
def ccp_verification(pdf_path, ccp_numbers):
    nb_attestations = convert_pdf_to_images(pdf_path)//2
    eel.sleep(0.01)
    if stop or nb_attestations == 0:
        eel.stop_process()
        return None

    for nb_att in range(1,nb_attestations+1):
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        #Faire appel à l'extraction de texte de l'attestation et du relevé.
        response_releve, image_releve = ocr_call('releve_'+ str(nb_att) +'.jpg')
        response_attestation, image_attestation = ocr_call('attestation_'+ str(nb_att) +'.jpg')
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        # Récupère et organise le résultat d'extraction.
        content_1, _= get_content(response_releve, image_releve)
        content_2, content_3 = get_content(response_attestation, image_attestation)
        lines_1, lines_dict_1 =  get_lines(content_1)
        lines_2, lines_dict_2 =  get_lines(content_2)
        lines_3, lines_dict_3 =  get_lines(content_3)

        ccp_numbers_1 = find_ccp_releve(is_flipped(image_releve, lines_dict_1, content_1), lines_dict_1, content_1)
        ccp_numbers_2 = find_ccp(lines_dict_2, content_2, image_attestation)
        ccp_numbers_3 = find_ccp(lines_dict_3, content_3, image_attestation)
        
        # Vérifiez si le système a trouvé le numéro de compte (CCP) dans le relevé
        found = 0
        for num in ccp_numbers_1:
            if ccp_numbers.get(int(num)) == 0:
                ccp_numbers[int(num)] = 1
                found = 1
                break
            
        # Vérifiez si le système a trouvé le numéro de CCP dans la partie gauche de l'attestation.
        if not found:
            for num in ccp_numbers_2.values():
                for i in range(1,10):
                    if ccp_numbers.get(int(num[:i])) == 0:
                        ccp_numbers[int(num[:i])] = 1
                        found = 1
        
        # Vérifiez si le système a trouvé le numéro de CCP dans la partie droite de l'attestation.
        if not found:
            for num in ccp_numbers_3.values():
                for i in range(1,10):
                    if ccp_numbers.get(int(num[:i])) == 0:
                        ccp_numbers[int(num[:i])] = 1
    return ccp_numbers

@eel.expose
def main(excel_path ,pdf_path, nom_page, colonne):
    try :
        reinitialize_stop()
        # Lecture des fichiers Excel et pdf
        output_path  = '/'.join(excel_path.split('/')[:-1])
        file_name = excel_path.split('/')[-1].split('.')[0]
        file_extention = excel_path.split('/')[-1].split('.')[-1]
        
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        
        df = pd.read_excel(excel_path, sheet_name=nom_page)
        
        # Vérification de l'existence des colonnes concernées.
        column_compte = colonne
        for column in df.columns:
            if column.lower().strip() == column_compte.lower().strip():
                column_compte = column
                break  
        if not column_compte:
            return f'Verifier le fichier excel si il a la column {column_compte}'
        
        column_nom = ''
        column_prenom = ''
        for column in df.columns:
            if column.lower().strip() == 'nom':
                column_nom = column 
            if column.lower().strip() == 'prenom':
                column_prenom = column
        
        # Vérifiez s'il y a un numéro de compte qui n'est pas valide.
        index_ccp_pb = []
        for value in zip(df[column_compte].values, df[column_compte].index):
            if value[0] is np.nan or len(re.findall('[^0-9]',str(value[0]).strip())) > 0:
                index_ccp_pb.append(value[1])
                
        index_ccp_pb = pd.Index(index_ccp_pb)
        df_ccp = df.drop(index=index_ccp_pb)
        ccp_numbers = dict(zip(df_ccp[column_compte], np.zeros(len(df_ccp[column_compte]))))
        ccp_numbers = OrderedDict(sorted(ccp_numbers.items(), key=lambda x:x[0],  reverse=True))
        
        # Attendre 10 secondes après le lancement de la fonctionnalité si l'utilisateur souhaite quitter l'opération.
        debut = time.time()
        while (time.time()-debut) < 10:
            eel.sleep(0.01)
            if stop:
                eel.stop_process()
                return None
        # Commencer les operations
        ccp_numbers = ccp_verification(pdf_path, ccp_numbers)
        
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        
        # Ajoutez la colonne "status" dans le fichier Excel.
        data = {column_compte: list(ccp_numbers.keys()), 'STATUS': list(ccp_numbers.values())}
        df_status = pd.DataFrame.from_dict(data)
        df_croisement = pd.merge(df,df_status, left_on=column_compte, right_on=column_compte, how='left')
        df_croisement['STATUS'] = df_croisement['STATUS'].replace({0:'<i class="fa-solid fa-circle-xmark" style="color: red;"></i>', 1: True})
        df_croisement['STATUS'] = df_croisement['STATUS'].fillna('<i class="fa-solid fa-circle-xmark" style="color: red;"></i>')
        df_croisemenet = df_croisement[[column_compte, column_nom, column_prenom, 'STATUS']]
        
        # Récupérer les comptes qui ne sont pas vérifiés.
        df_final = df_croisemenet[df_croisemenet['STATUS']!= True]
        false_indexes = df_final.index
        df_croisemenet = df_croisemenet.drop(columns=['STATUS'])
        df_croisement.index = df_croisement.index + 2
        df_final.index = df_final.index+2
        df_final = df_final.reset_index().rename(columns={"index": "LIGNE EXCEL"})
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        df_final = df_final.to_html(index=False)
        df_final = df_final.replace('text-align: right', 'text-align: center')
        df_final = df_final.replace('&lt;i class="fa-solid fa-circle-xmark" style="color: red;"&gt;&lt;/i&gt;', '<i class="fa-solid fa-circle-xmark" style="color: red;"></i>')
        df_final = df_final.replace('<tbody>', '<tbody style="text-align: center;">')
        df.to_excel(file_name+' coloré.'+file_extention, index = False)
        
        # Coloriez les lignes qui ne sont pas vérifiées.
        wb = load_workbook(file_name+' coloré.'+file_extention)
        ws = wb['Sheet1']
        color_rows(ws, len(df_croisement.columns), false_indexes , 'ff3333')
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        wb.save(file_name+' coloré.'+file_extention)
        eel.close_loading_popup(output_path, file_name+' coloré.'+file_extention, df_final) #chemin du fichier de retour
    
    except Exception as e:
        traceback_str = traceback.format_exc()
        with open('logs.txt', 'a') as file:
            file.write(f"\n\nError in ocr_call: {traceback_str}")
        eel.gestion_exception()

import time

@eel.expose
def division_excel (path, page_name = 'BATCH', line_numbers = 971,sheet_name='Feuil1'):
    try :
        
        global stop
        
        reinitialize_stop()
        line_numbers = int(line_numbers)
        page_name = str(page_name)
        df = pd.read_excel(path, sheet_name)
        if stop:
            eel.stop_process()
            return None
        
        # Convertir les numéros de comptes en chaînes de caractères.
        column_rib = ''
        column_compte = ''
        for column in df.columns:
            if column.lower().strip() == 'rib':
                column_rib = column 
            if column.lower().strip() == 'compte':
                column_compte = column
                
        if column_rib:
            df[column_rib] = df[column_rib].astype('str')
            df[column_rib] = df[column_rib].str.pad(width=20, fillchar='0')
        if column_compte:
            df[column_compte] = df[column_compte].astype('str')
            df[column_compte] = df[column_compte].str.pad(width=10, fillchar='0')
        
        column_compte_ccp = ''
        for column in df.columns:
            if column.lower().strip() == 'ccp client':
                column_compte_ccp = column
                # print(column)
                break  
        if column_compte_ccp:
            df[column_compte_ccp] = df[column_compte_ccp].astype('str')
            df[column_compte_ccp] = df[column_compte_ccp].str.pad(width=10, fillchar='0')
        
        totals = []
        file_path = '/'.join(path.split('/')[:-1])
        file_name = path.split('/')[-1].split('.')[0]
        file_extention = path.split('/')[-1].split('.')[-1]
        
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        column_montant_echeance = ''
        for column in df.columns:
            if column.lower().strip() == 'montant echeance':
                column_montant_echeance = column
                # print(column)
                break  
        if not column_montant_echeance:
            raise Exception("Column montant don't exist")
        
        with pd.ExcelWriter(file_name+' divisé.'+file_extention) as writer:
            df_vide = pd.DataFrame()
            df_vide.to_excel(writer, sheet_name='Total', index=False)
            sheet_num = 0
            for sheet_num in range(math.floor(df.shape[0] / line_numbers))[:-1]:
                
                eel.sleep(0.01)
                if stop:
                    eel.stop_process()
                    return None
                
                # Création de la feuille
                # Récupération des lignes
                start_idx = sheet_num * line_numbers
                end_idx = (sheet_num + 1) * line_numbers
                sheet_name = page_name+f'{sheet_num + 1}'
                df_slice = df.iloc[start_idx:end_idx]
                df_slice.index = range(1, len(df_slice) + 1)
                df_slice.rename_axis('N°', inplace=True)
                df_slice.reset_index(inplace=True)
                
                # Ajouter la somme des montants
                total = df_slice[column_montant_echeance].sum()
                new_row = pd.DataFrame({column_montant_echeance: [total]})
                df_slice = pd.concat([df_slice, new_row], ignore_index=True)
                df_slice.to_excel(writer, sheet_name=sheet_name, index=False)
                totals.append([total, df_slice.shape[0]-1])
            if not sheet_num:
                sheet_num =-1
            
            sheet_name = page_name+f'{sheet_num + 2}'
            start_idx = (sheet_num + 1) * line_numbers
            end_idx = (sheet_num + 2) * line_numbers
            eel.sleep(0.01)
            if stop:
                eel.stop_process()
                return None
            
            # Vérifier si la dernière feuille contient moins de 997 lignes
            # Si c'est le cas, on va créer une seule feuille 
            if 997-((df.shape[0] % line_numbers) + line_numbers+1) > 0:
    #           create_sheet(writer, df, sheet_name, start_idx, end_idx:df.shape[0])
                df_slice = df.iloc[start_idx:df.shape[0]]
                df_slice.index = range(1, len(df_slice) + 1)
                df_slice.rename_axis('N°', inplace=True)
                df_slice.reset_index(inplace=True)
                total = df_slice[column_montant_echeance].sum()
                new_row = pd.DataFrame({column_montant_echeance: [total]})
                df_slice = pd.concat([df_slice, new_row], ignore_index=True)
                df_slice.to_excel(writer, sheet_name=sheet_name, index=False)
                #if sheet_num+1 != df.shape[0] // 971:
                totals.append([total, df_slice.shape[0]-1])
                # Écrire la somme dans la cellule
            
            # Sinon, on va créer deux feuilles. La première a le numéro spécifié par le client,
            # et la deuxième feuille contient le reste.
            else:
                start_idx = (sheet_num+1) * line_numbers
                # print(start_idx, end_idx, df.shape[0])
    #             create_sheet(writer, df, sheet_name, start_idx, end_idx)
                df_slice = df.iloc[start_idx:end_idx]
                df_slice.index = range(1, len(df_slice) + 1)
                df_slice.rename_axis('N°', inplace=True)
                df_slice.reset_index(inplace=True)
                total = df_slice[column_montant_echeance].sum()
                new_row = pd.DataFrame({column_montant_echeance: [total]})
                df_slice = pd.concat([df_slice, new_row], ignore_index=True)
                df_slice.to_excel(writer, sheet_name=sheet_name, index=False)
                #if sheet_num+1 != df.shape[0] // 971:
                totals.append([total, df_slice.shape[0]-1])
                # Écrire la somme dans la cellule
                sheet_name = page_name+f'{sheet_num + 3}'
                df_slice = df.iloc[end_idx:df.shape[0]]
                df_slice.index = range(1, len(df_slice) + 1)
                df_slice.rename_axis('N°', inplace=True)
                df_slice.reset_index(inplace=True)
                total = df_slice['MONTANT ECHEANCE'].sum()
                new_row = pd.DataFrame({'MONTANT ECHEANCE': [total]})
                df_slice = pd.concat([df_slice, new_row], ignore_index=True)
                df_slice.to_excel(writer, sheet_name=sheet_name, index=False)
                totals.append([total, df_slice.shape[0]-1])
            # Crée une feuille qui contient le total.
            eel.sleep(0.01)
            if stop:
                eel.stop_process()
                return None
            batch_name = []
            totals = np.array(totals)
            totals = np.append(totals, [[sum(totals[:,0]),sum(totals[:,1])]], axis =0)
            for i in range(1,len(totals)):
                batch_name.append(page_name+f'{i}')
            batch_name.append('Total')
            df_total = pd.DataFrame(list(zip(batch_name, totals[:,0], totals[:,1])), columns =['Batch','Montant echeance total','Nombre de comptes'])
            df_total.index = range(1, df_total.shape[0]+1)
            df_total['Nombre de comptes'] = df_total['Nombre de comptes'].astype('int32')
            # Crée une feuille qui contient le total.
            df_total.to_excel(writer, sheet_name='Total', index=False)
                
        eel.close_loading_popup(file_path, file_name+' divisé.'+file_extention) #chemin du fichier de retour
    except :
        eel.gestion_exception()

# coloré les lignes
def color_rows(ws, nb_columns, indexes, color):
    global stop
    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for i in indexes:
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        for col in alphabet[:nb_columns]:
            ws[col+str(i+2)].fill = PatternFill(patternType='solid',fgColor=color)

# verification de la situation 
@eel.expose
def verification_compte(excel_path, nom_feuille1, colonne1, situation_compte_path, nom_feuille2, colonne2):
    try:
        reinitialize_stop()
        global stop
        # lecture des fichiers excel
        output_path = '/'.join(excel_path.split('/')[:-1])
        file_name = excel_path.split('/')[-1].split('.')[0]
        file_extention = excel_path.split('/')[-1].split('.')[-1]

        df_croise = pd.read_excel(excel_path, nom_feuille1)
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        
        df_situation = pd.read_excel(situation_compte_path, nom_feuille2)    

        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        
        # Verification des colonnes
        column_compte = ''
        for column in df_croise.columns:
            if column.lower().strip() == colonne1:
                column_compte = column
                break
        column_nb_compte = ''
        for column in df_situation.columns:
            if column.lower().strip() == colonne2:
                column_nb_compte = column
                break  
        
        if not column_nb_compte or not column_compte:
            raise Exception("Column Compte don't exist")
        
        df_situation = df_situation[[column_nb_compte]]
        # Vérifiez s'il y a un problème dans les colonnes.
        index_rib_pb = []
        for value in zip(df_croise[column_compte].values, df_croise[column_compte].index):
            if value[0] is np.nan or len(re.findall('[^0-9]',str(value[0]).strip())) > 0:
                index_rib_pb.append(value[1])
        index_rib_pb = pd.Index(index_rib_pb)
        df_rib_pb = df_croise.iloc[index_rib_pb]
        df_rib_pb.loc[:,['Etat']] = 'PB RIB'
        
        df_croise.drop(index=index_rib_pb, inplace=True)
        df_croise = df_croise.astype({column_compte: 'int64'})
        df_situation = df_situation.astype({column_nb_compte: 'int64'})
        
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        # Croisez les fichiers en fonction du numéro de compte.
        df_croisement = pd.merge(df_croise,df_situation, left_on=column_compte, right_on=column_nb_compte, how='left')
        df_croisement['Etat'] = np.where((df_croisement[column_nb_compte].isna()),'NON ACTIF','ACTIF')
        df_croisement.drop(columns = [column_nb_compte], inplace = True)
        df_croisement.reset_index(drop=True,inplace=True)
        df_croisement = pd.concat([df_croisement, df_rib_pb], axis=0)
        
        # Convertir les colonnes de compte en chaîne de caractères.
        column_rib = ''
        column_ccp = ''
        column_ccp_asba = ''
        for column in df_croisement.columns:
            if column.lower().strip() == 'rib':
                column_rib = column
            if column.lower().strip() == 'ccp client':
                column_ccp = column
            if column.lower().strip() == 'ccp asba':
                column_ccp_asba = column
        if column_rib:
            df_croisement[column_rib] = df_croisement[column_rib].astype('str')
            df_croisement[column_rib] = df_croisement[column_rib].str.pad(width=20, fillchar='0')
            
        if column_ccp:
            df_croisement[column_ccp] = df_croisement[column_ccp].astype('str')
            df_croisement[column_ccp] = df_croisement[column_ccp].str.pad(width=10, fillchar='0')

        #df_croisement.sort_index(inplace=True)
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        
        if column_ccp_asba:
            df_croisement[column_ccp_asba] = df_croisement[column_ccp_asba].astype('str')
            df_croisement[column_ccp_asba] = df_croisement[column_ccp_asba].str.pad(width=10, fillchar='0')
        
        if column_compte:
            df_croisement = df_croisement.astype({column_compte: 'str'})
            df_croisement[column_compte] = df_croisement[column_compte].str.pad(width=10, fillchar='0')
        # Créez le fichier Excel résultat
        # df_croisement.sort_index(inplace=True)
        df_croisement.to_excel(file_name+' vérifié.'+file_extention, index=False)
        wb = load_workbook(file_name+' vérifié.'+file_extention)
        ws = wb['Sheet1']
        indexes = df_croisement[df_croisement['Etat'] == 'NON ACTIF'].index
        color_rows(ws, len(df_croisement.columns), indexes, 'ff3333')
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        
        indexes = range(df_croise.shape[0], df_croise.shape[0]+len(df_croisement[df_croisement['Etat'] == 'PB RIB']))
        # Coloriez les lignes non actives.
        color_rows(ws, len(df_croisement.columns), indexes , 'f5933d')
        eel.sleep(0.01)
        if stop:
            eel.stop_process()
            return None
        wb.save(file_name+' vérifié.'+file_extention)

        eel.close_loading_popup(output_path, file_name+' vérifié.'+file_extention) #chemin du fichier de retour
    except Exception as e:
        eel.gestion_exception()

@eel.expose
def close_process(*args):
    global stop
    stop = True


@eel.expose
def reinitialize_stop(*args):
    global stop
    stop = False
    


@eel.expose
def close_python(*args):
    # Lorsque vous avez fini, libérez le verrou en fermant le fichier
    # Récupérer le chemin du répertoire actuel
    repertoire_actuel = os.getcwd()

    # Liste tous les fichiers dans le répertoire actuel
    fichiers_dans_repertoire = os.listdir(repertoire_actuel)

    # Parcourir la liste des fichiers et supprimer les fichiers Excel
    for fichier in fichiers_dans_repertoire:
        if fichier.endswith(".xlsx") or fichier.endswith(".jpg"):
            chemin_fichier = os.path.join(repertoire_actuel, fichier)
            os.remove(chemin_fichier)
    #os.remove(lock_file)
    sys.exit()



def log_exception(exc_type, exc_value, exc_traceback):
    with open('error.log', 'a') as f:
        f.write(f"Exception Type: {exc_type}\n")
        f.write(f"Exception Value: {exc_value}\n")
        f.write(f"Traceback:\n")
        traceback.print_tb(exc_traceback, file=f)

sys.excepthook = log_exception

if __name__ == '__main__':
    
    global stop
    stop = False
    eel.start("index.html", size=(1366, 743), port=0, block=True)